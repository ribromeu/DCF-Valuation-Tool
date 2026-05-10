"""
╔══════════════════════════════════════════════════════════════════╗
║              DCF VALUATION TOOL  —  v3.0                        ║
║                                                                  ║
║  ENTRY POINT:                                                    ║
║    python dcf_valuation.py                                       ║
║                                                                  ║
║  MODELS:                                                         ║
║    1. Traditional DCF        — discounted FCF + Gordon TV        ║
║    2. Gordon (Dividend DDM)  — P = D1 / (Ke − g)                ║
║    3. Gordon (FCF-based TV)  — TV = FCF*(1+g) / (WACC−g)        ║
║    4. Sector Multiple        — EV = Multiple × Metric            ║
║    5. Sensitivity Matrix     — 9×8 grid anchored to base WACC/g  ║
║                                                                  ║
║  EXCEL:                                                          ║
║    All computed cells use live =FORMULA references to            ║
║    Inputs Log so changing one number cascades everywhere.        ║
║                                                                  ║
║  REQUIREMENTS:                                                   ║
║    pip install openpyxl pandas numpy                             ║
╚══════════════════════════════════════════════════════════════════╝
"""

# ──────────────────────────────────────────────────────────────────
# BLOCK 0 — IMPORTS AND GLOBAL CONFIGURATION
#
# WHAT IT DOES: imports all required libraries and defines the global
# color palette for the Excel workbook. Every sheet builder reads
# from these constants — a full retheme is a one-place change.
#
# DEPENDENCIES: none (entry point)
# USED IN: all subsequent blocks
# ──────────────────────────────────────────────────────────────────

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from datetime import datetime
import os, io

C_DARK   = "1A1A2E"   # page background / header rows
C_MID    = "16213E"   # even data rows
C_ACCENT = "0F3460"   # odd data rows / section headers
C_GREEN  = "00B894"   # upside / buy signal
C_RED    = "D63031"   # downside / sell signal
C_YELLOW = "FDCB6E"   # neutral zone in sensitivity heatmap
C_PURPLE = "A29BFE"   # sector multiple accent
C_BLUE   = "0984E3"   # Gordon model accent
C_WHITE  = "FFFFFF"   # primary text
C_GRAY   = "B2BEC3"   # secondary text / footnotes


# ──────────────────────────────────────────────────────────────────
# BLOCK 1 — EXCEL STYLE HELPERS
#
# WHAT IT DOES: thin wrappers that stamp the dark-navy theme onto any
# cell — fill, font, border, alignment in one call. Also provides
# mh() (merge + header) and ref() (absolute cell address string).
#
# WHY THIS WAY: openpyxl requires a brand-new style object per cell;
# reusing the same instance silently corrupts the second assignment.
# These helpers construct fresh objects on every call intentionally.
# ref() centralises the "$A$1" string format so formula strings
# built in BLOCK 8 are readable and consistent.
#
# DEPENDENCIES: BLOCK 0 (color constants)
# USED IN: BLOCK 8 (all sheet builders)
# ──────────────────────────────────────────────────────────────────

def _fill(h):
    return PatternFill("solid", fgColor=h)

def _font(bold=False, color=C_WHITE, size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _border():
    s = Side(style="thin", color="2D3436")
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def sc(cell, value, bg=C_MID, bold=False, color=C_WHITE):
    """Apply full dark-theme style to a single cell."""
    cell.value     = value
    cell.fill      = _fill(bg)
    cell.font      = _font(bold=bold, color=color)
    cell.border    = _border()
    cell.alignment = _center()

def mh(ws, row, c1, c2, text, bg=C_ACCENT, size=12):
    """Merge a row range and write a bold section header."""
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row=row, column=c1, value=text)
    c.fill      = _fill(bg)
    c.font      = _font(bold=True, size=size)
    c.alignment = _center()

def ref(sheet, row, col):
    """Return an absolute cross-sheet cell reference: ='Sheet'!$C$5"""
    return f"='{sheet}'!${get_column_letter(col)}${row}"


# ──────────────────────────────────────────────────────────────────
# BLOCK 2 — INPUT HELPERS
#
# WHAT IT DOES: three validated wrappers around input():
#   • ask()              — numeric value with type enforcement + default
#   • ask_str()          — string value with default
#   • ask_fcfs_by_year() — interactive year-by-year FCF collection;
#                          each prompt is labeled with the calendar
#                          year so year–value mapping is unambiguous
#   • section()          — prints a labeled divider between input blocks
#
# WHY THIS WAY: validation loops prevent silent crashes from bad input
# (e.g. "10.5%" instead of "10.5"). Named-year prompts eliminate the
# off-by-one confusion of comma-separated lists ("is year 3 the third
# or fourth number?"). Defaults let users press Enter through standard
# cases without retyping.
#
# DEPENDENCIES: none
# USED IN: BLOCK 9 (main — all six input sections)
# ──────────────────────────────────────────────────────────────────

def ask(prompt, type_=float, default=None):
    suffix = f" [{default}]" if default is not None else ""
    while True:
        raw = input(f"  {prompt}{suffix}: ").strip()
        if raw == "" and default is not None:
            return default
        try:
            return type_(raw)
        except ValueError:
            print(f"    ⚠  Invalid value. Expected: {type_.__name__}")

def ask_str(prompt, default=None):
    suffix = f" [{default}]" if default is not None else ""
    raw = input(f"  {prompt}{suffix}: ").strip()
    return raw if raw else (default or "")

def ask_fcfs_by_year(start_year):
    # Repeats the horizon question until the user gives a valid integer
    # in [1, 30] — an empty list or a 31-year horizon are both caught
    while True:
        n = ask("Number of forecast years", type_=int, default=5)
        if 1 <= n <= 30:
            break
        print("    ⚠  Enter a value between 1 and 30.")
    print()
    fcfs, years = [], []
    for i in range(n):
        yr  = start_year + i
        val = ask(f"  FCF  Year {yr}", default=0.0)
        fcfs.append(val)
        years.append(yr)
    return fcfs, years

def section(title):
    print(f"\n{'─'*54}")
    print(f"  {title}")
    print(f"{'─'*54}")


# ──────────────────────────────────────────────────────────────────
# BLOCK 3 — TRADITIONAL DCF MODEL
#
# WHAT IT DOES: multi-period discounted cash flow valuation using
# a Gordon Growth perpetuity as the terminal value:
#
#   PV(FCF_t) = FCF_t / (1 + WACC)^t          t = 1 … n
#   TV        = FCF_n × (1 + g) / (WACC − g)  Gordon perpetuity
#   PV(TV)    = TV / (1 + WACC)^n
#   EV        = Σ PV(FCF_t) + PV(TV)
#   Price     = (EV − Net Debt) / Shares
#
# WHY THIS WAY: Gordon TV is the industry standard for perpetuity-
# based DCF (Damodaran, 2012). The model is single-path and
# deterministic; uncertainty is handled separately in BLOCK 6
# (sensitivity matrix). WACC > g is a hard pre-condition —
# enforced in BLOCK 9 before this function is ever called.
#
# DEPENDENCIES: none
# USED IN: BLOCK 6 (sensitivity), BLOCK 7 (terminal report),
#          BLOCK 8 (Excel — Summary, FCF Projections)
# ──────────────────────────────────────────────────────────────────

def dcf_model(fcfs, wacc, g_terminal, shares, net_debt):
    pv_fcfs = [fcf / (1 + wacc)**i for i, fcf in enumerate(fcfs, 1)]
    tv      = fcfs[-1] * (1 + g_terminal) / (wacc - g_terminal)
    pv_tv   = tv / (1 + wacc)**len(fcfs)
    ev      = sum(pv_fcfs) + pv_tv
    equity  = ev - net_debt
    price   = equity / shares if shares > 0 else 0
    return {
        "pv_fcfs":           pv_fcfs,
        "sum_pv_fcfs":       sum(pv_fcfs),
        "terminal_value":    tv,
        "pv_terminal_value": pv_tv,
        "enterprise_value":  ev,
        "equity_value":      equity,
        "price_per_share":   price,
    }


# ──────────────────────────────────────────────────────────────────
# BLOCK 4 — GORDON GROWTH MODEL (TWO VARIANTS)
#
# WHAT IT DOES: estimates intrinsic value via two complementary
# Gordon Growth formulations:
#
#   Variant A — Dividend DDM (equity-level):
#     P = D1 / (Ke − g)
#     Discounts dividends using the cost of equity. Applicable when
#     the firm pays dividends and Ke > g.
#
#   Variant B — FCF-based TV (firm-level):
#     TV = FCF_last × (1 + g) / (WACC − g)
#     Equity = (TV − Net Debt) / Shares
#     Uses the last projected FCF as the perpetuity base. This is
#     exactly the terminal value mechanism inside the DCF model, but
#     presented as a standalone output — useful to isolate how much
#     of the DCF price is driven by the TV assumption alone.
#
# WHY TWO VARIANTS: the dividend DDM and FCF-TV Gordon models answer
# different questions. DDM answers "what is this dividend stream
# worth to an equity holder?". FCF-TV answers "if the firm grew its
# cash flows at g forever from year n, what is the firm worth today?"
# A large divergence between the two signals either a payout policy
# gap (retained earnings not reinvested productively) or a mismatch
# between the FCF projections and dividend capacity.
#
# EDGE CASES: both return None when the denominator ≤ 0.
#
# DEPENDENCIES: none
# USED IN: BLOCK 7 (terminal report), BLOCK 8 (Excel — Summary)
# ──────────────────────────────────────────────────────────────────

def gordon_dividend(d1, ke, g):
    """Variant A: P = D1 / (Ke − g)"""
    if ke <= g:
        return None
    return d1 / (ke - g)

def gordon_fcf(fcfs, wacc, g, net_debt, shares):
    """Variant B: standalone FCF perpetuity — same math as DCF terminal value."""
    if wacc <= g:
        return None
    tv     = fcfs[-1] * (1 + g) / (wacc - g)
    equity = tv - net_debt
    return equity / shares if shares > 0 else 0


# ──────────────────────────────────────────────────────────────────
# BLOCK 5 — SECTOR MULTIPLE MODEL
#
# WHAT IT DOES: relative valuation via a market-derived multiple:
#
#   EV     = Multiple × Metric
#   Equity = EV − Net Debt
#   Price  = Equity / Shares
#
# WHY THIS WAY: multiples anchor value to what the market currently
# pays for comparable firms. Not a substitute for DCF — a sanity
# check. A large DCF/multiples divergence usually means either the
# FCF projections are aggressive or the sector multiple is stale.
# EV/EBITDA is the most common cross-sector multiple because it is
# capital-structure neutral (pre-interest, pre-taxes, pre-D&A).
#
# DEPENDENCIES: none
# USED IN: BLOCK 7 (terminal report), BLOCK 8 (Excel — Summary)
# ──────────────────────────────────────────────────────────────────

def multiples_model(metric, multiple, net_debt, shares):
    ev     = multiple * metric
    equity = ev - net_debt
    price  = equity / shares if shares > 0 else 0
    return {"enterprise_value": ev, "equity_value": equity, "price_per_share": price}


# ──────────────────────────────────────────────────────────────────
# BLOCK 6 — SENSITIVITY ANALYSIS
#
# WHAT IT DOES: builds a 9 × 8 matrix of DCF implied share prices
# by varying WACC and terminal g around the user's base-case inputs.
#
#   WACC range: base ± 2pp in 8 equally spaced steps
#   g range:    base ± 2pp in 9 equally spaced steps
#   (steps are clamped so neither axis goes below 0.1% or above 30%)
#
# WHY 9×8 AND ±2pp: the previous ±3–7pp range produced extreme
# values that were visually noisy and rarely actionable. A ±2pp band
# covers the realistic estimation uncertainty of WACC (beta
# estimation error ≈ ±1–2pp per Damodaran) while keeping the base
# case visible at the center of the matrix. The 9×8 grid (72 cells)
# is large enough to show convexity without overflowing a screen.
#
# BASE CASE HIGHLIGHT: the cell corresponding exactly to (wacc, g)
# is tagged in the returned metadata so the Excel builder can apply
# a distinct border to identify the actual base-case assumption.
#
# EDGE CASE: cells where WACC ≤ g remain None (model undefined).
#
# DEPENDENCIES: BLOCK 3 (dcf_model)
# USED IN: BLOCK 7 (terminal report), BLOCK 8 (Excel — Sensitivity)
# ──────────────────────────────────────────────────────────────────

def sensitivity_analysis(fcfs, net_debt, shares, wacc_base, g_base):
    # Fixed shock steps applied symmetrically around the base-case value.
    # Each entry is a delta in percentage points added to the base rate:
    #   ±0.01, ±0.02, ±0.03, ±0.04, ±0.05, ±0.10  plus the base (0)
    # This produces a 13×13 matrix — fine-grained nearby moves on the
    # inner rows/cols, a wide stress scenario on the outer ±10% band.
    #
    # WHY FIXED STEPS instead of linspace: fixed increments make the
    # headers self-explanatory ("−1%" / "+5%") regardless of what the
    # user's base WACC or g happens to be. The reader always knows
    # exactly what shock each column represents.
    #
    # np.clip keeps every resulting rate in (0.001, 0.49) to prevent
    # undefined or economically nonsensical values.
    shocks = [-0.10, -0.05, -0.04, -0.03, -0.02, -0.01,
               0.00,
               0.01,  0.02,  0.03,  0.04,  0.05,  0.10]

    wacc_range = np.clip([wacc_base + s for s in shocks], 0.001, 0.49)
    g_range    = np.clip([g_base    + s for s in shocks], 0.001, 0.48)

    rows = []
    for g in g_range:
        row = []
        for w in wacc_range:
            if w <= g:
                row.append(None)
            else:
                r = dcf_model(fcfs, w, g, shares, net_debt)
                row.append(round(r["price_per_share"], 2))
        rows.append(row)

    # Labels show the shock, not the absolute rate — clearer for presentation
    shock_labels = ["-10%","-5%","-4%","-3%","-2%","-1%",
                    "base",
                    "+1%","+2%","+3%","+4%","+5%","+10%"]

    df = pd.DataFrame(rows, index=shock_labels, columns=shock_labels)

    # Base case is always the center cell (position 6 in 13 elements)
    base_row = 6
    base_col = 6

    return df, base_row, base_col


# ──────────────────────────────────────────────────────────────────
# BLOCK 7 — TERMINAL REPORT
#
# WHAT IT DOES: prints a structured summary to the terminal:
#   • Section 1: model comparison (price, upside, signal) — now
#     includes both Gordon variants as separate rows
#   • Section 2: DCF breakdown (PV FCFs, TV, EV, net debt, equity)
#   • Section 3: sensitivity matrix in fixed-width grid
#
# WHY THIS WAY: terminal output is a fast sanity check before the
# user opens the Excel file. Fixed-width formatting (no tabulate)
# keeps the dependency list minimal.
#
# DEPENDENCIES: BLOCKS 3, 4, 5, 6
# USED IN: BLOCK 9 (main)
# ──────────────────────────────────────────────────────────────────

def print_results(inp, dcf, gp_div, gp_fcf, mt, sens_df):
    cp  = inp["current_price"]
    cur = inp["currency"]
    W   = 60

    print(f"\n{'═'*W}")
    print(f"  RESULTS  —  {inp['company_name']} ({inp['ticker']})")
    print(f"{'═'*W}")

    # ── Section 1: model comparison
    print(f"\n  {'MODEL':<32} {'PRICE':>10}  {'UPSIDE':>8}  SIGNAL")
    print(f"  {'─'*W}")
    for label, price in [
        ("Traditional DCF",         dcf["price_per_share"]),
        ("Gordon — Dividend DDM",   gp_div),
        ("Gordon — FCF Terminal",   gp_fcf),
        ("Sector Multiple",         mt["price_per_share"]),
    ]:
        if price is None:
            print(f"  {label:<32} {'N/A':>10}  {'—':>8}  undefined")
        else:
            u = (price - cp) / cp
            s = "▲ BUY" if u > 0.10 else ("▼ SELL" if u < -0.10 else "◆ NEUTRAL")
            print(f"  {label:<32} {cur} {price:>7.2f}  {u*100:>+7.1f}%  {s}")
    print(f"  {'─'*W}")
    print(f"  {'Market Price':<32} {cur} {cp:>7.2f}")

    # ── Section 2: DCF breakdown
    tv_pct = (dcf["pv_terminal_value"] / dcf["enterprise_value"] * 100
              if dcf["enterprise_value"] != 0 else 0.0)
    print(f"\n  DCF BREAKDOWN")
    print(f"  {'─'*W}")
    for i, pv in enumerate(dcf["pv_fcfs"], 1):
        print(f"  {'PV FCF Year ' + str(i):<32} {pv:>14,.2f}")
    for lbl, val in [
        ("PV of Projected FCFs",  dcf["sum_pv_fcfs"]),
        ("PV of Terminal Value",  dcf["pv_terminal_value"]),
        ("Enterprise Value (EV)", dcf["enterprise_value"]),
        ("(-) Net Debt",          inp["net_debt"]),
        ("Equity Value",          dcf["equity_value"]),
        ("Implied Price (DCF)",   dcf["price_per_share"]),
    ]:
        print(f"  {lbl:<32} {val:>14,.2f}")
    print(f"  {'TV / EV':<32} {tv_pct:>13.1f}%")

    # ── Section 3: sensitivity matrix
    print(f"\n  SENSITIVITY  (DCF price — WACC vs. g)   base: "
          f"WACC={inp['wacc']*100:.2f}%  g={inp['g_terminal']*100:.2f}%")
    print(f"  {'WACC →':<10}", end="")
    for col in sens_df.columns:
        print(f"{col:>9}", end="")
    print()
    print(f"  {'─'*W}")
    for idx, row in sens_df.iterrows():
        print(f"  g={idx:<8}", end="")
        for val in row:
            print(f"{'      N/A' if val is None else f'{val:>9.2f}'}", end="")
        print()
    print()


# ──────────────────────────────────────────────────────────────────
# BLOCK 8 — EXCEL WORKBOOK BUILDER
#
# WHAT IT DOES: constructs a four-sheet Excel workbook and returns
# raw bytes. The key design constraint for v3.0 is maximum formula
# referencing — every numeric cell that can be derived from another
# cell uses a live =FORMULA instead of a hardcoded string.
#
# FORMULA STRATEGY:
#   • Inputs Log (Sheet 4) holds all raw numeric inputs as true
#     numbers (not formatted strings). It is the single source of
#     truth — like a named-range parameter table.
#   • Summary, FCF Projections, and Sensitivity cells reference
#     Inputs Log via ='Inputs Log'!$B$n. Changing any input there
#     cascades to all three analytical sheets automatically.
#   • FCF Projections computes Discount Factor, PV, cumulative PV,
#     and FCF/Share entirely with =FORMULAS, so the sheet is a live
#     model — the user can override FCF values in Excel and all
#     downstream cells update immediately.
#   • Summary valuation rows compute price and upside via formulas
#     referencing both the FCF sheet and Inputs Log.
#
# WHY RETURNING BYTES: keeps the function pure — the caller decides
# the file path. Makes the function testable without disk I/O.
#
# DEPENDENCIES: BLOCK 1 (style helpers), BLOCKS 3–6 (result dicts)
# USED IN: BLOCK 9 (main)
# ──────────────────────────────────────────────────────────────────

def build_excel(inp, dcf, gp_div, gp_fcf, mt, sens_df,
                base_row, base_col, fcfs):
    wb  = Workbook()
    cp  = inp["current_price"]
    cur = inp["currency"]
    n   = len(fcfs)           # number of forecast years
    IL  = "Inputs Log"        # sheet name used in cross-references

    # ══════════════════════════════════════════════════════════════
    # SHEET 4 BUILT FIRST — Inputs Log is the formula source of truth.
    # All other sheets reference rows in this sheet. Build it now so
    # the row numbers are known when we write the formulas below.
    # ══════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet(IL)
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 34
    ws4.column_dimensions["B"].width = 20
    ws4.column_dimensions["C"].width = 34
    ws4.column_dimensions["D"].width = 20

    mh(ws4, 2, 1, 4, "FULL INPUT LOG  —  source of truth for all formulas", C_DARK, 13)
    for j, h in enumerate(["Parameter", "Value", "Parameter", "Value"], 1):
        sc(ws4.cell(4, j), h, C_ACCENT, bold=True)

    # Numeric inputs stored as raw numbers (no formatting) so Excel
    # formulas can do arithmetic on them directly. Text fields go in
    # col A/B only and are not referenced by formulas.
    #
    # Row map — used to build cross-sheet references:
    IL_ROW = {}   # name → row number in Inputs Log

    left_params = [
        ("wacc",        inp["wacc"]),
        ("ke",          inp["ke"]),
        ("kd",          inp["kd"]),
        ("tax",         inp["tax"]),
        ("g_terminal",  inp["g_terminal"]),
        ("g_gordon",    inp["g_gordon"]),
        ("net_debt",    inp["net_debt"]),
        ("shares",      inp["shares"]),
        ("current_price", inp["current_price"]),
        ("d1",          inp["d1"]),
        ("multiple",    inp["multiple"]),
        ("metric",      inp["metric"]),
    ]
    right_params = [
        ("Company",    inp["company_name"]),
        ("Ticker",     inp["ticker"]),
        ("Sector",     inp["sector"]),
        ("Currency",   inp["currency"]),
        ("Mult. Type", inp["multiple_type"]),
        ("# Years",    n),
        ("Timestamp",  datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    # FCF rows appended after the fixed params
    fcf_params = [(f"fcf_{i+1}", v) for i, v in enumerate(fcfs)]

    all_left = left_params + fcf_params
    for i, (name, val) in enumerate(all_left, 5):
        row = i
        IL_ROW[name] = row
        ws4.row_dimensions[row].height = 20
        bg  = C_MID if i % 2 == 0 else C_ACCENT
        bg2 = C_MID if i % 2 == 0 else C_DARK
        sc(ws4.cell(row, 1), name, bg,  bold=True)
        # Store raw number so formulas can reference it
        cell = ws4.cell(row, 2)
        cell.value     = val
        cell.fill      = _fill(bg2)
        cell.font      = _font()
        cell.border    = _border()
        cell.alignment = _center()

    for i, (name, val) in enumerate(right_params, 5):
        row = i
        ws4.row_dimensions[row].height = 20
        bg  = C_MID if i % 2 == 0 else C_ACCENT
        bg2 = C_MID if i % 2 == 0 else C_DARK
        sc(ws4.cell(row, 3), name, bg,  bold=True)
        sc(ws4.cell(row, 4), str(val), bg2)

    # ── helper: absolute reference to a named input in Inputs Log
    def ir(name):
        """='Inputs Log'!$B$row  for the named parameter."""
        return f"='{IL}'!$B${IL_ROW[name]}"

    # ══════════════════════════════════════════════════════════════
    # SHEET 2: FCF Projections — fully formula-driven
    # ══════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("FCF Projections")
    ws2.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFG", [8, 24, 18, 24, 18, 24, 22]):
        ws2.column_dimensions[col].width = w

    mh(ws2, 2, 1, 7, "FREE CASH FLOW PROJECTIONS", C_DARK, 13)
    for j, h in enumerate(["Year", "Projected FCF", "Discount Factor",
                             "PV of FCF", "% of Total PV",
                             "Cumulative PV", "FCF / Share"], 1):
        sc(ws2.cell(4, j), h, C_ACCENT, bold=True)

    # Row where each year's data lives in ws2 (used for cumulative formula)
    fcf_data_rows = []
    for i in range(1, n + 1):
        row  = 4 + i
        bg   = C_MID if i % 2 == 0 else C_ACCENT
        fcf_data_rows.append(row)
        ws2.row_dimensions[row].height = 22

        B = get_column_letter(2)   # FCF column
        C = get_column_letter(3)   # Discount Factor column
        D = get_column_letter(4)   # PV column
        E = get_column_letter(5)   # % column
        F = get_column_letter(6)   # Cumulative column
        G = get_column_letter(7)   # FCF/Share column

        fcf_ref    = f"'{IL}'!$B${IL_ROW[f'fcf_{i}']}"   # raw FCF from Inputs Log
        wacc_ref   = f"'{IL}'!$B${IL_ROW['wacc']}"
        shares_ref = f"'{IL}'!$B${IL_ROW['shares']}"

        # Year label (ordinal)
        sc(ws2.cell(row, 1), i, bg)

        # Projected FCF — references Inputs Log directly
        c = ws2.cell(row, 2)
        c.value     = f"={fcf_ref}"
        c.fill      = _fill(bg); c.font = _font(); c.border = _border(); c.alignment = _center()

        # Discount Factor = 1 / (1 + WACC)^t
        c = ws2.cell(row, 3)
        c.value     = f"=1/(1+{wacc_ref})^{i}"
        c.fill      = _fill(bg); c.font = _font(); c.border = _border(); c.alignment = _center()

        # PV of FCF = FCF × Discount Factor  (B × C, both in same row)
        c = ws2.cell(row, 4)
        c.value     = f"={B}{row}*{C}{row}"
        c.fill      = _fill(bg); c.font = _font(); c.border = _border(); c.alignment = _center()

        # % of Total PV — SUM of D column for all years computed after loop;
        # reference will be filled in as a named range placeholder for now
        # using the total PV row that will be written after the loop.
        # We write a placeholder and fix it below.
        c = ws2.cell(row, 5)
        c.value     = f"={D}{row}/SUM({D}5:{D}{4+n})"
        c.number_format = "0.0%"
        c.fill      = _fill(bg); c.font = _font(); c.border = _border(); c.alignment = _center()

        # Cumulative PV = SUM of D5 : D{row}
        c = ws2.cell(row, 6)
        c.value     = f"=SUM({D}5:{D}{row})"
        c.fill      = _fill(bg); c.font = _font(); c.border = _border(); c.alignment = _center()

        # FCF / Share
        c = ws2.cell(row, 7)
        c.value     = f"={B}{row}/{shares_ref}"
        c.fill      = _fill(bg); c.font = _font(); c.border = _border(); c.alignment = _center()

    # Terminal Value row
    tv_row = 4 + n + 1
    ws2.row_dimensions[tv_row].height = 22
    g_ref    = f"'{IL}'!$B${IL_ROW['g_terminal']}"
    wacc_ref = f"'{IL}'!$B${IL_ROW['wacc']}"
    nd_ref   = f"'{IL}'!$B${IL_ROW['net_debt']}"
    sh_ref   = f"'{IL}'!$B${IL_ROW['shares']}"
    last_fcf = f"B{4+n}"       # last projected FCF row in ws2

    # TV = FCF_n * (1+g) / (WACC - g)
    tv_formula   = f"={last_fcf}*(1+{g_ref})/({wacc_ref}-{g_ref})"
    # PV(TV) = TV / (1+WACC)^n
    pv_tv_formula = f"=C{tv_row}/(1+{wacc_ref})^{n}"

    sc(ws2.cell(tv_row, 1), "TV",  C_DARK, bold=True)
    # TV raw value
    c = ws2.cell(tv_row, 2)
    c.value = tv_formula
    c.fill = _fill(C_DARK); c.font = _font(bold=True); c.border = _border(); c.alignment = _center()
    # Discount factor for TV = 1/(1+WACC)^n
    c = ws2.cell(tv_row, 3)
    c.value = f"=1/(1+{wacc_ref})^{n}"
    c.fill = _fill(C_DARK); c.font = _font(); c.border = _border(); c.alignment = _center()
    # PV(TV)
    c = ws2.cell(tv_row, 4)
    c.value = pv_tv_formula
    c.fill = _fill(C_DARK); c.font = _font(bold=True); c.border = _border(); c.alignment = _center()
    # % of total (PV FCFs + PV TV)
    total_ev_formula = f"SUM(D5:D{4+n})+D{tv_row}"
    c = ws2.cell(tv_row, 5)
    c.value = f"=D{tv_row}/({total_ev_formula})"
    c.number_format = "0.0%"
    c.fill = _fill(C_DARK); c.font = _font(); c.border = _border(); c.alignment = _center()
    # Enterprise Value = sum PV FCFs + PV TV
    c = ws2.cell(tv_row, 6)
    c.value = f"=SUM(D5:D{4+n})+D{tv_row}"
    c.fill = _fill(C_DARK); c.font = _font(bold=True); c.border = _border(); c.alignment = _center()
    sc(ws2.cell(tv_row, 7), "—", C_DARK)

    # Summary totals row
    tot_row = tv_row + 1
    ws2.row_dimensions[tot_row].height = 22
    sc(ws2.cell(tot_row, 1), "EV",  C_ACCENT, bold=True)
    for j, formula in enumerate([
        f"=SUM(B5:B{4+n})",                 # total projected FCF
        "—",
        f"=SUM(D5:D{tv_row})",              # total PV (FCFs + TV)
        "=100%",
        f"=F{tv_row}",                       # EV (same as cumulative after TV)
        "—",
    ], 2):
        c = ws2.cell(tot_row, j)
        c.value     = formula
        c.fill      = _fill(C_ACCENT); c.font = _font(bold=True)
        c.border    = _border(); c.alignment = _center()



    # ══════════════════════════════════════════════════════════════
    # SHEET 1: Summary — references FCF Projections and Inputs Log
    # ══════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([32, 24, 32, 24, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 38

    mh(ws, 2, 1, 5,
       f"DCF VALUATION REPORT  —  {inp['company_name']} ({inp['ticker']})  |  {inp['sector']}",
       C_DARK, 14)

    # Inputs block — labels in col A/C, live formula values in col B/D
    r = 4
    mh(ws, r, 1, 4, "INPUTS", C_ACCENT)
    r += 1

    def fmt_pct(name):
        """Formula that reads a decimal rate and displays as percentage."""
        return f"={ir(name)[1:]}*100&\"%\""   # strip leading = for concat

    input_rows = [
        ("WACC",                ir("wacc"),          "Cost of Equity (Ke)", ir("ke")),
        ("Cost of Debt (Kd)",   ir("kd"),            "Tax Rate",            ir("tax")),
        ("Terminal Growth g",   ir("g_terminal"),    "Gordon Growth g",     ir("g_gordon")),
        ("Net Debt",            ir("net_debt"),      "Shares Outstanding",  ir("shares")),
        ("Sector Multiple",     ir("multiple"),      "Metric Value",        ir("metric")),
        ("Next Dividend D1",    ir("d1"),            "Current Market Price",ir("current_price")),
    ]
    for l1, v1, l2, v2 in input_rows:
        ws.row_dimensions[r].height = 22
        sc(ws.cell(r,1), l1, C_ACCENT, bold=True)
        c = ws.cell(r,2); c.value = v1
        c.fill=_fill(C_MID); c.font=_font(); c.border=_border(); c.alignment=_center()
        sc(ws.cell(r,3), l2, C_ACCENT, bold=True)
        c = ws.cell(r,4); c.value = v2
        c.fill=_fill(C_MID); c.font=_font(); c.border=_border(); c.alignment=_center()
        r += 1

    # Valuation results
    r += 1
    mh(ws, r, 1, 5, "VALUATION RESULTS", C_ACCENT)
    r += 1
    for h, c in zip(["Model","Intrinsic Price","Market Price","Upside / Downside","Signal"], range(1,6)):
        sc(ws.cell(r,c), h, C_DARK, bold=True)
    r += 1

    # Market price reference for upside formula
    mkt = f"'{IL}'!$B${IL_ROW['current_price']}"

    # FCF Projections sheet references for Gordon FCF and DCF
    ev_ref    = f"='FCF Projections'!F{tv_row}"    # Enterprise Value
    pv_tv_ref = f"='FCF Projections'!D{tv_row}"    # PV Terminal Value
    pv_fcf_ref= f"='FCF Projections'!D{tot_row}"   # total PV of all FCFs+TV

    model_rows = []
    for (name, price_py, price_formula, color) in [
        ("Traditional DCF",
         dcf["price_per_share"],
         f"=({ev_ref}-{nd_ref})/{sh_ref}",
         C_GREEN),
        ("Gordon — Dividend DDM",
         gp_div,
         (f"={ir('d1')[1:]}/({ir('ke')[1:]}-{ir('g_gordon')[1:]})"
          if gp_div is not None else None),
         C_BLUE),
        ("Gordon — FCF Terminal",
         gp_fcf,
         (f"=('FCF Projections'!B{4+n}*(1+{ir('g_terminal')[1:]})"
          f"/({ir('wacc')[1:]}-{ir('g_terminal')[1:]})-{ir('net_debt')[1:]})"
          f"/{ir('shares')[1:]}"
          if gp_fcf is not None else None),
         C_PURPLE),
        ("Sector Multiple",
         mt["price_per_share"],
         f"=({ir('multiple')[1:]}*{ir('metric')[1:]}-{ir('net_debt')[1:]})/{ir('shares')[1:]}",
         C_YELLOW),
    ]:
        ws.row_dimensions[r].height = 24
        price_val = price_py   # Python-computed for color decisions
        if price_val is None or price_formula is None:
            for col in range(1,6):
                sc(ws.cell(r,col), "N/A  (undefined)", C_MID)
        else:
            upside_py = (price_val - cp) / cp
            bg_up  = C_GREEN if upside_py > 0 else C_RED
            signal = "BUY" if upside_py > 0.10 else ("SELL" if upside_py < -0.10 else "NEUTRAL")

            sc(ws.cell(r,1), name, C_MID)

            # Intrinsic price — live formula
            c = ws.cell(r,2)
            c.value = price_formula
            c.fill=_fill(C_MID); c.font=_font(); c.border=_border(); c.alignment=_center()

            # Market price — live reference
            c = ws.cell(r,3)
            c.value = f"={mkt}"
            c.fill=_fill(C_MID); c.font=_font(); c.border=_border(); c.alignment=_center()

            # Upside = (B_r - C_r) / C_r, formatted as %
            c = ws.cell(r,4)
            c.value = f"=(B{r}-C{r})/C{r}"
            c.number_format = "+0.0%;-0.0%;0.0%"
            c.fill=_fill(bg_up); c.font=_font(bold=True); c.border=_border(); c.alignment=_center()

            sc(ws.cell(r,5), signal, bg_up, bold=True)
        model_rows.append(r)
        r += 1

    # DCF Breakdown — all formulas referencing FCF Projections sheet
    r += 1
    mh(ws, r, 1, 4, "DCF BREAKDOWN", C_ACCENT)
    r += 1

    sum_pv_row   = tot_row     # EV row in FCF Projections (col D = total PV)
    breakdown = [
        ("PV of Projected FCFs",   f"=SUM('FCF Projections'!D5:'FCF Projections'!D{4+n})"),
        ("Terminal Value (TV)",    f"='FCF Projections'!B{tv_row}"),
        ("PV of Terminal Value",   f"='FCF Projections'!D{tv_row}"),
        ("Enterprise Value (EV)",  f"='FCF Projections'!F{tv_row}"),
        ("(-) Net Debt",           ir("net_debt")),
        ("Equity Value",           f"='FCF Projections'!F{tv_row}-{ir('net_debt')[1:]}"),
        ("Implied Price (DCF)",    f"=(='FCF Projections'!F{tv_row}-{ir('net_debt')[1:]})/{ir('shares')[1:]}"),
        ("TV / EV",                f"='FCF Projections'!D{tv_row}/'FCF Projections'!F{tv_row}"),
    ]
    # Fix the Implied Price formula (extra = inside string)
    breakdown[6] = ("Implied Price (DCF)",
                    f"=('FCF Projections'!F{tv_row}-{ir('net_debt')[1:]})/{ir('shares')[1:]}")
    tv_pct_row = None
    for lbl, formula in breakdown:
        ws.row_dimensions[r].height = 22
        sc(ws.cell(r,1), lbl, C_ACCENT, bold=True)
        c = ws.cell(r,2)
        c.value = formula
        if lbl == "TV / EV":
            c.number_format = "0.0%"
            tv_pct_row = r
        c.fill=_fill(C_MID); c.font=_font(); c.border=_border(); c.alignment=_center()
        ws.cell(r,3).fill = _fill(C_DARK)
        ws.cell(r,4).fill = _fill(C_DARK)
        r += 1

    # Timestamp
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(r, 1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    c.fill=_fill(C_DARK); c.font=_font(color=C_GRAY, size=9); c.alignment=_center()

    # ══════════════════════════════════════════════════════════════
    # SHEET 3: Sensitivity — color-coded 9×8 heatmap
    # Python-computed values (can't formula-drive a variable-size
    # matrix in openpyxl without VBA), but market-price reference
    # row shows a live formula for the current base-case price.
    # ══════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Sensitivity")
    ws3.sheet_view.showGridLines = False
    nc = len(sens_df.columns)   # 8 WACC values
    nr = len(sens_df.index)     # 9 g values

    mh(ws3, 1, 1, nc+2,
       "SENSITIVITY ANALYSIS  —  Implied Price  (WACC  ×  terminal g)   base highlighted",
       C_DARK, 13)
    ws3.row_dimensions[1].height = 30

    legend_items = [
        (C_GREEN,  "Upside > 20%"),
        ("55EFC4", "Upside 5–20%"),
        (C_YELLOW, "Upside ±5%"),
        ("E17055", "Downside 5–20%"),
        (C_RED,    "Downside > 20%"),
    ]
    for j, (bg, lbl) in enumerate(legend_items, 1):
        c = ws3.cell(2, j)
        c.value     = lbl
        c.fill      = _fill(bg)
        c.font      = _font(size=9, color=C_DARK if bg in ("55EFC4", C_YELLOW) else C_WHITE)
        c.alignment = _center()
        c.border    = _border()

    # Row 3: market price (live formula) and WACC column headers
    sc(ws3.cell(3, 1), "Market Price →", C_DARK, bold=True)
    c = ws3.cell(3, 2)
    c.value = f"={mkt}"     # live reference to current market price
    c.fill=_fill(C_DARK); c.font=_font(bold=True, color=C_YELLOW)
    c.border=_border(); c.alignment=_center()
    for j, col in enumerate(sens_df.columns, 3):
        sc(ws3.cell(3, j), col, C_ACCENT, bold=True)

    # Data rows — Python values with color encoding; base-case cell
    # gets a thick green border to mark the actual assumption used
    for i, (idx, row_data) in enumerate(sens_df.iterrows(), 4):
        sc(ws3.cell(i, 1), "", C_DARK)
        sc(ws3.cell(i, 2), idx, C_ACCENT, bold=True)
        for j, val in enumerate(row_data, 3):
            is_base = (i - 4 == base_row) and (j - 3 == base_col)
            if val is None:
                sc(ws3.cell(i, j), "N/A", C_DARK, color=C_GRAY)
            else:
                upside = (val - cp) / cp if cp else 0
                if upside > 0.20:    bg, txt = C_GREEN,  C_WHITE
                elif upside > 0.05:  bg, txt = "55EFC4", C_DARK
                elif upside > -0.05: bg, txt = C_YELLOW, C_DARK
                elif upside > -0.20: bg, txt = "E17055", C_WHITE
                else:                bg, txt = C_RED,    C_WHITE
                c = ws3.cell(i, j)
                c.value     = val
                c.fill      = _fill(bg)
                c.font      = _font(color=txt, size=11,
                                    bold=is_base)   # bold = base case
                c.alignment = _center()
                # Thick border on base-case cell to make it findable
                if is_base:
                    thick = Side(style="medium", color=C_GREEN)
                    c.border = Border(left=thick, right=thick,
                                      top=thick,  bottom=thick)
                else:
                    c.border = _border()

    for col in range(1, nc + 4):
        ws3.column_dimensions[get_column_letter(col)].width = 12

    # Re-order sheets: Summary first
    wb.move_sheet("Summary",      offset=-(wb.index(ws)))
    wb.move_sheet("FCF Projections", offset=1 - wb.index(wb["FCF Projections"]) + wb.index(ws))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────
# BLOCK 9 — MAIN
#
# WHAT IT DOES: orchestrates the full valuation workflow:
#   1. Six input sections (company → capital → structure → FCF →
#      income models → sensitivity range)
#   2. Two pre-condition checks (WACC > g, non-zero FCFs)
#   3. All five models + sensitivity grid
#   4. Terminal report
#   5. Excel export
#
# WHY THIS WAY: the six-section flow mirrors how an analyst actually
# builds a model — identity first, then cost of capital, then
# structure, then cash flows, then income-based cross-checks, then
# stress tests. The order prevents inputs from being entered before
# their dependencies are defined.
#
# DEPENDENCIES: all previous blocks
# USED IN: script entry point (__name__ == "__main__")
# ──────────────────────────────────────────────────────────────────

def main():
    print()
    print("╔══════════════════════════════════════════════════════╗")
    print("║           DCF VALUATION TOOL  —  v3.0               ║")
    print("║  DCF  ·  Gordon (DDM + FCF)  ·  Sector Multiple     ║")
    print("║  Sensitivity  ·  Excel with live formulas            ║")
    print("╚══════════════════════════════════════════════════════╝")

    section("1 / 6   COMPANY IDENTIFICATION")
    company  = ask_str("Company name",  "MyCompany")
    ticker   = ask_str("Ticker",        "TICK")
    sector   = ask_str("Sector",        "Energy")
    currency = ask_str("Currency",      "USD")

    # WACC blends debt and equity costs weighted by capital structure.
    # Ke is used exclusively by the Gordon DDM (equity discount rate).
    section("2 / 6   COST OF CAPITAL")
    wacc = ask("WACC (%)",              default=10.5) / 100
    ke   = ask("Cost of Equity Ke (%)", default=12.0) / 100
    kd   = ask("Cost of Debt Kd (%)",   default=8.0)  / 100
    tax  = ask("Tax rate (%)",           default=21.0) / 100

    section("3 / 6   CAPITAL STRUCTURE")
    net_debt = ask("Net Debt",             default=50_000_000.0)
    shares   = ask("Shares outstanding",   default=10_000_000.0)
    cp       = ask("Current market price", default=35.0)

    # FCFs entered year-by-year to prevent positional ambiguity.
    # g_terminal must be < WACC — checked immediately after entry.
    section("4 / 6   FREE CASH FLOW  &  TERMINAL GROWTH")
    start_year = ask("Forecast start year", type_=int,
                     default=datetime.now().year + 1)
    print()
    fcfs, years = ask_fcfs_by_year(start_year)
    g_terminal  = ask("Terminal growth rate g (%)", default=3.0) / 100

    if wacc <= g_terminal:
        print(f"\n  ⚠  WACC ({wacc*100:.2f}%) must be > g ({g_terminal*100:.2f}%). Rerun.\n")
        return
    if all(f == 0 for f in fcfs):
        print("\n  ⚠  All FCFs are zero. Enter at least one non-zero value.\n")
        return

    # Gordon DDM uses D1 and Ke (equity-level).
    # Gordon FCF uses the last projected FCF and WACC (firm-level).
    section("5 / 6   GORDON GROWTH  &  SECTOR MULTIPLE")
    d1       = ask("Next year dividend D1",      default=2.5)
    g_gordon = ask("Gordon DDM growth rate g (%)", default=4.0) / 100

    print()
    print("  Multiple types: EV/EBITDA, P/E, EV/EBIT, EV/Revenue, P/Book")
    mult_type = ask_str("Multiple type",                     "EV/EBITDA")
    multiple  = ask("Multiple value",                        default=8.5)
    metric    = ask("Metric value (EBITDA, Earnings, ...)",  default=20_000_000.0)

    # Sensitivity anchored ±2pp around the base-case WACC and g.
    # No manual range input needed — the grid centers automatically.
    section("6 / 6   SENSITIVITY  (auto-centered ±2pp around base)")
    print(f"  Base WACC: {wacc*100:.2f}%  →  range [{(wacc-0.02)*100:.2f}%, {(wacc+0.02)*100:.2f}%]")
    print(f"  Base g:    {g_terminal*100:.2f}%  →  range [{(g_terminal-0.02)*100:.2f}%, {(g_terminal+0.02)*100:.2f}%]")
    print(f"  Grid: 9 g values × 8 WACC values  =  72 cells")

    print(f"\n  {'─'*52}")
    print("  Running calculations...")

    dcf    = dcf_model(fcfs, wacc, g_terminal, shares, net_debt)
    gp_div = gordon_dividend(d1, ke, g_gordon)
    gp_fcf = gordon_fcf(fcfs, wacc, g_terminal, net_debt, shares)
    mt     = multiples_model(metric, multiple, net_debt, shares)
    sens_df, base_row, base_col = sensitivity_analysis(
        fcfs, net_debt, shares, wacc, g_terminal)

    inp = dict(
        company_name=company, ticker=ticker, sector=sector, currency=currency,
        wacc=wacc, ke=ke, kd=kd, tax=tax,
        g_terminal=g_terminal, g_gordon=g_gordon,
        net_debt=net_debt, shares=shares, current_price=cp,
        d1=d1, multiple=multiple, multiple_type=mult_type, metric=metric,
        forecast_years=years,
    )

    print_results(inp, dcf, gp_div, gp_fcf, mt, sens_df)

    fname = f"DCF_{ticker}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    xl    = build_excel(inp, dcf, gp_div, gp_fcf, mt,
                        sens_df, base_row, base_col, fcfs)
    with open(fname, "wb") as f:
        f.write(xl)

    print(f"  ✅  Excel saved: {fname}")
    print(f"  📁  Path: {os.path.abspath(fname)}")
    print()


if __name__ == "__main__":
    main()
